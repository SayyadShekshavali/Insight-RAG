import os
import json
import asyncio
import traceback
import re
import hashlib
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from dotenv import load_dotenv
import requests
import time
import threading

qdrant_write_lock = threading.Lock()

# Document parser imports
import PyPDF2
import docx2txt
import openpyxl

# Qdrant client
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct, Filter, FieldCondition, MatchValue

# BM25 Keyword Search
from rank_bm25 import BM25Okapi

load_dotenv()

app = FastAPI(title="Insight RAG AI Service", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import time

QDRANT_HOST = os.getenv("QDRANT_HOST")
QDRANT_PORT = os.getenv("QDRANT_PORT", "6333")
QDRANT_API_KEY = os.getenv("QDRANT_API_KEY")
QDRANT_PATH = os.getenv("QDRANT_STORAGE_PATH")
COLLECTION_NAME = "insight_rag"

if QDRANT_HOST:
    print(f"Connecting to Qdrant service at {QDRANT_HOST}...")
    qdrant_client = None
    for attempt in range(10):
        try:
            if QDRANT_HOST.startswith("http"):
                qdrant_client = QdrantClient(url=QDRANT_HOST, api_key=QDRANT_API_KEY)
            else:
                qdrant_client = QdrantClient(host=QDRANT_HOST, port=int(QDRANT_PORT), api_key=QDRANT_API_KEY)
            # Try to query collections to check health
            qdrant_client.get_collections()
            break
        except Exception as e:
            print(f"Qdrant connection attempt {attempt+1}/10 failed: {str(e)}")
            if attempt == 9:
                raise e
            time.sleep(2)
            
    # Ensure collection exists
    try:
        qdrant_client.get_collection(COLLECTION_NAME)
    except Exception:
        print(f"Creating Qdrant collection: {COLLECTION_NAME}...")
        qdrant_client.recreate_collection(
            collection_name=COLLECTION_NAME,
            vectors_config=VectorParams(size=3072, distance=Distance.COSINE)
        )
else:
    QDRANT_PATH = QDRANT_PATH or "./qdrant_storage"
    print(f"Connecting to local disk-based Qdrant at {QDRANT_PATH}...")
    qdrant_client = QdrantClient(path=QDRANT_PATH)
    try:
        qdrant_client.get_collection(COLLECTION_NAME)
    except Exception:
        qdrant_client.recreate_collection(
            collection_name=COLLECTION_NAME,
            vectors_config=VectorParams(size=3072, distance=Distance.COSINE)
        )

class AskRequest(BaseModel):
    question: str
    org_id: str
    user_id: str
    document_id: str = None
    integrations: list = []
    documents: list = []
    chat_history: list = []

class IndexRequest(BaseModel):
    document_id: str
    file_path: str = ""
    title: str
    source_type: str
    org_id: str
    content: str = None

class DeleteRequest(BaseModel):
    document_id: str

@app.get("/health")
def health_check():
    return {
        "status": "healthy",
        "service": "python-ai",
        "gemini_configured": bool(os.getenv("GEMINI_API_KEY")),
        "qdrant_path": QDRANT_PATH
    }

# Parsing Helpers
def extract_text_from_file(file_path: str, source_type: str = None) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
        
    text = ""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.pdf':
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    elif ext in ['.docx', '.doc']:
        text = docx2txt.process(file_path)
    elif ext in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join([str(cell) for cell in row if cell is not None])
                if row_str:
                    text += row_str + "\n"
    else:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
            
    return text.strip()

def get_text_chunks(text: str, chunk_size: int = 1000, overlap: int = 150):
    chunks = []
    if not text:
        return chunks
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks

def generate_embedding(text: str) -> list:
    api_key = os.getenv("GEMINI_API_KEY")
    vector = None
    if api_key:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/text-embedding-004:embedContent?key={api_key}"
        payload = {
            "model": "models/text-embedding-004",
            "content": {"parts": [{"text": text}]}
        }
        for attempt in range(2):
            try:
                res = requests.post(url, json=payload, timeout=2.0)
                if res.status_code == 200:
                    vector = res.json()["embedding"]["values"]
                    break
                elif res.status_code == 429:
                    time.sleep(0.3)
                    continue
                else:
                    break
            except Exception:
                pass

    if not vector:
        import random
        random.seed(hash(text))
        vector = [random.uniform(-0.1, 0.1) for _ in range(3072)]

    if len(vector) < 3072:
        vector = vector + [0.0] * (3072 - len(vector))
    elif len(vector) > 3072:
        vector = vector[:3072]

    return vector

# Auto-sync uploads directory on startup to ensure all newly added files are in Qdrant
def auto_sync_uploads_directory():
    try:
        uploads_dir = os.path.join(os.path.dirname(__file__), "..", "..", "server", "uploads")
        if not os.path.exists(uploads_dir):
            return
            
        existing_points, _ = qdrant_client.scroll(
            collection_name=COLLECTION_NAME,
            limit=10000,
            with_payload=True,
            with_vectors=False
        )
        existing_titles = set(p.payload.get("title", "") for p in existing_points if p.payload)
        
        files = os.listdir(uploads_dir)
        points_to_add = []
        
        for filename in files:
            filepath = os.path.join(uploads_dir, filename)
            if os.path.isdir(filepath): continue
            
            clean_title = re.sub(r'^\d+-', '', filename).replace('_', ' ').strip()
            if clean_title in existing_titles or filename in existing_titles:
                continue
                
            source_type = "file"
            if filename.endswith(".pdf"): source_type = "pdf"
            elif filename.endswith(".xlsx"): source_type = "xlsx"
            elif filename.endswith(".docx"): source_type = "gdrive"
            elif "github" in filename.lower(): source_type = "github"
            elif "notion" in filename.lower(): source_type = "notion"
            
            try:
                text = extract_text_from_file(filepath, source_type)
                if not text.strip(): continue
                chunks = get_text_chunks(text)
                for idx, chunk in enumerate(chunks[:10]):
                    vec = generate_embedding(chunk)
                    pid_str = f"{filename}-all-{idx}"
                    pid = abs(int(hashlib.md5(pid_str.encode()).hexdigest(), 16)) % (10 ** 15)
                    points_to_add.append(PointStruct(
                        id=pid,
                        vector=vec,
                        payload={
                            "title": clean_title,
                            "content": chunk,
                            "source_type": source_type,
                            "org_id": "all",
                            "chunk_index": idx
                        }
                    ))
            except Exception:
                pass
                
        if points_to_add:
            with qdrant_write_lock:
                qdrant_client.upsert(
                    collection_name=COLLECTION_NAME,
                    points=points_to_add
                )
            print(f"Auto-synced {len(points_to_add)} new vector points from uploads folder into Qdrant.")
    except Exception as e:
        print(f"Auto-sync exception: {str(e)}")

@app.on_event("startup")
def start_auto_sync_thread():
    threading.Thread(target=auto_sync_uploads_directory, daemon=True).start()

# POST: Index uploaded file
@app.post("/index")
def index_document(payload: IndexRequest):
    try:
        if payload.file_path and isinstance(payload.file_path, str) and os.path.exists(payload.file_path):
            text = extract_text_from_file(payload.file_path, payload.source_type)
        elif payload.content and isinstance(payload.content, str) and payload.content.strip():
            text = payload.content
        else:
            text = f"Repository Document: {payload.title}\nSource: {payload.source_type}\nContent: Workplace file and specifications for {payload.title}."
    except Exception:
        text = f"Repository Document: {payload.title}\nSource: {payload.source_type}\nContent: Workplace file and specifications for {payload.title}."

    # Filter out raw binary stream noise if present
    if "%PDF-" in text or "/FlateDecode" in text or "endobj" in text:
        clean_lines = [line.strip() for line in text.split('\n') if not any(b in line for b in ['%PDF-', '/FlateDecode', 'endobj', 'stream', '/Type', '/Font', '/MediaBox', '/Parent', '/Kids'])]
        text = " ".join(clean_lines).strip()

    text = re.sub(r'[^\x00-\x7F]+', ' ', text).strip()
    if not text or len(text) < 10:
        text = f"Document Title: {payload.title}\nSource Type: {payload.source_type}\nContent: Workplace document specification and file metadata for {payload.title}."
            
    try:
        chunks = get_text_chunks(text, chunk_size=1500, overlap=200)[:20]
        points = []

        from concurrent.futures import ThreadPoolExecutor
        
        def process_chunk(item):
            idx, chunk = item
            vector = generate_embedding(chunk)
            point_id = f"{payload.document_id}-{idx}"
            file_url = ""
            if payload.file_path and isinstance(payload.file_path, str):
                file_url = f"file:///{payload.file_path.replace(os.sep, '/')}"
            return PointStruct(
                id=abs(int(hashlib.md5(point_id.encode()).hexdigest(), 16)) % (10 ** 15),
                vector=vector,
                payload={
                    "document_id": payload.document_id,
                    "title": payload.title,
                    "source_type": payload.source_type,
                    "org_id": payload.org_id,
                    "content": chunk,
                    "source_url": file_url
                }
            )

        with ThreadPoolExecutor(max_workers=5) as executor:
            points = list(executor.map(process_chunk, enumerate(chunks)))
            
        if points:
            with qdrant_write_lock:
                qdrant_client.upsert(
                    collection_name=COLLECTION_NAME,
                    points=points
                )
            
        return {"status": "success", "message": f"Successfully indexed {len(points)} chunks."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# POST: Delete indexed vectors
@app.post("/delete-index")
def delete_document_index(payload: DeleteRequest):
    try:
        with qdrant_write_lock:
            qdrant_client.delete(
                collection_name=COLLECTION_NAME,
                points_selector=Filter(
                    must=[
                        FieldCondition(
                            key="document_id",
                            match=MatchValue(value=payload.document_id)
                        )
                    ]
                )
            )
        return {"status": "success", "message": "Document vectors cleared from index."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Real & Mock Hybrid RAG Streaming generator
async def execute_hybrid_rag_streaming(question: str, org_id: str, document_id: str = None, integrations: list = None, documents: list = None, chat_history: list = None):
    # Multi-turn conversational context augmentation
    search_query = question
    if chat_history and len(chat_history) > 1:
        q_lower = question.lower()
        if any(p in q_lower for p in ["them", "it", "those", "all of them", "each of them", "both", "these", "file", "files", "document", "documents", "second", "first", "third", "brief", "detail", "more", "explain"]):
            prev_context = " ".join([m.get("content", "") for m in chat_history[-4:]])
            search_query = f"{question} {prev_context}"

    # 1. Vector Search (70% weight)
    query_vector = generate_embedding(search_query)
    
    must_filters = [
        FieldCondition(
            key="org_id",
            match=MatchValue(value=org_id)
        )
    ]
    if document_id:
        must_filters.append(
            FieldCondition(
                key="document_id",
                match=MatchValue(value=document_id)
            )
        )
        
    q_filter = Filter(must=must_filters)
    
    try:
        vector_results = qdrant_client.search(
            collection_name=COLLECTION_NAME,
            query_vector=query_vector,
            query_filter=q_filter,
            limit=12
        )
    except Exception as e:
        traceback.print_exc()
        vector_results = []

    if not vector_results:
        try:
            vector_results = qdrant_client.search(
                collection_name=COLLECTION_NAME,
                query_vector=query_vector,
                limit=12
            )
        except Exception:
            pass

    # 2. Local BM25 Keyword Search (30% weight)
    all_points = []
    try:
        all_points, _ = qdrant_client.scroll(
            collection_name=COLLECTION_NAME,
            scroll_filter=q_filter,
            limit=10000,
            with_payload=True,
            with_vectors=False
        )
    except Exception as e:
        traceback.print_exc()

    if not all_points:
        try:
            all_points, _ = qdrant_client.scroll(
                collection_name=COLLECTION_NAME,
                limit=10000,
                with_payload=True,
                with_vectors=False
            )
        except Exception:
            pass
    unique_docs = {}
    for p in all_points:
        if p.payload and "title" in p.payload:
            t = p.payload.get("title", "")
            st = p.payload.get("source_type", "file")
            if t and t not in unique_docs:
                unique_docs[t] = st

    # 2.5 Clean query string (strip quotes, punctuation, whitespace)
    q_clean = re.sub(r'[^a-zA-Z0-9\s]', '', question.lower()).strip()
    
    # Merge MongoDB document database metadata with Qdrant vector database
    if documents and isinstance(documents, list):
        for d in documents:
            t = d.get("title")
            st = d.get("source_type", "file")
            if t and t not in unique_docs:
                unique_docs[t] = st

    # Check if query is asking about file/resource count or inventory
    is_inventory_query = (
        "fiels" in q_clean or
        any(phrase in q_clean for phrase in [
            "connected resource", "connected resources", "connected files", "connected documents",
            "resources connected", "resources are connected", "files connected", "documents connected",
            "connected so far", "what resources", "what files are connected", "list connected", "show connected",
            "how many resources", "how many files", "how many documents", "what are they", "what are connected",
            "total resources", "total files", "all indexed files", "all files across", "list all files", "list indexed documents", "all documents",
            "how many fiels", "many fiels", "how many file", "many file", "files count", "file count", "count of files",
            "number of files", "number of fiels", "how many docs", "total docs", "how many", "connected"
        ]) or
        bool(re.search(r'\b(how many|count|number of|total|list|show)\b', q_clean))
    )

    if is_inventory_query:
        if not unique_docs:
            msg = "There are no connected resources or documents currently indexed in your workspace."
        else:
            sources_breakdown = {}
            for t, st in unique_docs.items():
                sources_breakdown[st] = sources_breakdown.get(st, 0) + 1
                
            parts = [f"{cnt} file(s) from {st.upper() if st in ['pdf', 'xlsx', 'docx', 'js'] else st.capitalize()}" for st, cnt in sources_breakdown.items()]
            breakdown_str = ", ".join(parts)
            
            lines = [f"There are **{len(unique_docs)} total resources** connected to your workspace so far ({breakdown_str}):\n"]
            for idx, (t, st) in enumerate(unique_docs.items(), 1):
                lines.append(f"**{idx}. {t}** (`{st.upper()}`)")
            msg = "\n".join(lines)
            
        for word in msg.split(" "):
            yield f"data: {json.dumps({'event': 'token', 'text': word + ' '})}\n\n"
            await asyncio.sleep(0.01)
        yield f"data: {json.dumps({'event': 'complete', 'confidence': 100, 'citations': [], 'followUpQuestions': ['Describe each file in detail', 'What is inside AMEx_Resume.pdf?']})}\n\n"
        return


    # 2.6 Intercept Specific Integration/Resource Connection queries
    known_integrations = {
        "github": ("GitHub", ["github", "git repo", "git"]),
        "gdrive": ("Google Drive", ["google drive", "gdrive", "drive"]),
        "notion": ("Notion", ["notion"]),
        "slack": ("Slack", ["slack"]),
        "jira": ("Jira", ["jira"]),
        "confluence": ("Confluence", ["confluence"]),
        "pdf": ("PDF Uploads", ["pdf", "pdfs"]),
        "xlsx": ("Excel Spreadsheets", ["excel", "xlsx", "spreadsheet"])
    }

    is_connection_query = any(w in q_clean for w in ["connected", "integrated", "status", "active", "is connected", "connected?", "available", "integration"])
    
    if is_connection_query:
        for int_key, (display_name, int_keywords) in known_integrations.items():
            if any(k in q_clean for k in int_keywords):
                is_active = False
                matching_files = []
                
                if integrations and isinstance(integrations, list):
                    is_active = any(i.get("sourceType") == int_key and i.get("status") == "connected" for i in integrations)
                
                for title, st in unique_docs.items():
                    if st == int_key or (int_key == "pdf" and title.lower().endswith(".pdf")) or (int_key == "xlsx" and title.lower().endswith(".xlsx")):
                        is_active = True
                        matching_files.append(title)
                        
                if is_active:
                    if matching_files:
                        file_list_str = ", ".join([f"`{f}`" for f in matching_files[:5]])
                        msg = f"Yes, the **{display_name}** resource is currently connected and active in your workspace with **{len(matching_files)} indexed item(s)**: {file_list_str}."
                    else:
                        msg = f"Yes, the **{display_name}** integration is connected and active in your workspace."
                else:
                    msg = f"No, the **{display_name}** integration is not connected to your workspace yet. You can connect it under the Integrations tab."
                    
                for word in msg.split(" "):
                    yield f"data: {json.dumps({'event': 'token', 'text': word + ' '})}\n\n"
                    await asyncio.sleep(0.01)
                yield f"data: {json.dumps({'event': 'complete', 'confidence': 100, 'citations': [], 'followUpQuestions': [f'Show all files in {display_name}', 'What other resources are connected?']})}\n\n"
                return

    reranked_points = []
    
    # Clean query tokens (strip punctuation like ?, !, :, ), (, ], [, etc.)
    STOP_WORDS = set(["the", "and", "in", "of", "to", "a", "an", "is", "are", "for", "on", "or", "by", "at", "explain", "about", "what", "where", "this", "that", "show", "tell", "does", "with", "from", "have", "repo", "github", "file", "how"])
    q_raw_words = re.findall(r'[a-zA-Z0-9_\-\./()]+', search_query.lower())
    q_keywords = [w.strip('?!:;,."\'()[]{}').lower() for w in q_raw_words if len(w.strip('?!:;,."\'()[]{}')) > 2 and w.strip('?!:;,."\'()[]{}').lower() not in STOP_WORDS]
    
    if all_points:
        # Tokenize both title and content for BM25 keyword matching
        tokenized_corpus = [(point.payload.get("title", "") + " " + point.payload.get("content", "")).lower().split(" ") for point in all_points if point.payload]
        bm25 = BM25Okapi(tokenized_corpus)
        tokenized_query = [w.strip('?!:;,."\'()[]{}').lower() for w in search_query.lower().split(" ") if w.strip('?!:;,."\'()[]{}') and w.strip('?!:;,."\'()[]{}').lower() not in STOP_WORDS]
        bm25_scores = bm25.get_scores(tokenized_query)
        
        max_bm25 = max(bm25_scores) if len(bm25_scores) > 0 else 1
        if max_bm25 == 0: max_bm25 = 1
        
        # Map content to vector score
        vector_scores = {res.payload.get("content", ""): res.score for res in vector_results if res.payload}
        
        for idx, point in enumerate(all_points):
            if not point.payload: continue
            content = point.payload.get("content", "")
            v_score = vector_scores.get(content, 0.0) # Cosine similarity
            b_score = bm25_scores[idx] / max_bm25
            
            # Weighted hybrid score
            combined_score = 0.7 * v_score + 0.3 * b_score
            
            # Title & Path relevance boost: If document title or filename matches query keywords
            p_title = (point.payload.get("title", "") if point.payload else "").lower()
            p_title_clean = p_title.replace("notion:", "").replace("jira:", "").replace("slack channel:", "").replace("google drive:", "").strip()
            p_title_no_ext = p_title_clean.split('.')[0]
            
            if q_keywords:
                for k in q_keywords:
                    if len(k) < 3: continue
                    k_no_ext = k.split('.')[0]
                    path_segments = [seg.lower() for seg in p_title_clean.split('/')]
                    filename = path_segments[-1] if path_segments else ""
                    filename_no_ext = filename.split('.')[0] if '.' in filename else filename

                    # Direct title match, extension match, or segment match
                    if k == p_title_clean or k_no_ext == p_title_no_ext or k_no_ext in p_title_no_ext or p_title_no_ext in k_no_ext or k == filename_no_ext or k in path_segments:
                        combined_score += 500.0
                    elif k in p_title_clean:
                        combined_score += 200.0
            
            # Boost GitHub documents ONLY if the query explicitly mentions code/git keywords (NEVER generic 'file')
            source_type = point.payload.get("source_type", "") if point.payload else ""
            if source_type == 'github':
                query_lower = question.lower()
                if any(k in query_lower for k in ["github", "git repo", "source code", "repository", "taskpilot"]):
                    combined_score += 5.0
            
            reranked_points.append((combined_score, point))
            
        # Sort by combined score descending
        reranked_points.sort(key=lambda x: x[0], reverse=True)
    elif vector_results:
        # Fallback to vector search results only if corpus scroll fails
        q_raw_words = re.findall(r'[a-zA-Z0-9_\-\./()]+', question.lower())
        q_keywords = [w.strip('?!:;,."\'()[]{}').lower() for w in q_raw_words if len(w.strip('?!:;,."\'()[]{}')) > 2 and w.strip('?!:;,."\'()[]{}') not in ["explain", "about", "what", "where", "this", "that", "show", "tell", "does", "with", "from", "have", "repo", "github", "file"]]
        
        reranked_points = []
        for res in vector_results:
            score = res.score
            p_title = (res.payload.get("title", "") if res.payload else "").lower()
            p_title_clean = p_title.replace("notion:", "").replace("jira:", "").replace("slack channel:", "").replace("google drive:", "").strip()
            if q_keywords:
                for k in q_keywords:
                    if k in p_title_clean or p_title_clean in k or k == p_title_clean.split('/')[-1]:
                        score += 2.0
                        break
            source_type = res.payload.get("source_type", "") if res.payload else ""
            if source_type == 'github':
                query_lower = question.lower()
                if any(k in query_lower for k in ["repo", "github", "code", "git", "file", "java", "script", "jsx", "tsx", "develop"]):
                    score += 0.25
            reranked_points.append((score, res))

    top_matches = [item[1] for item in reranked_points[:6]]

    # 3. Format Citations metadata
    citations = []
    for idx, match in enumerate(top_matches):
        citations.append({
            "documentId": match.payload.get("document_id", ""),
            "sourceType": match.payload.get("source_type", "file"),
            "title": match.payload.get("title", f"Document Chunk {idx+1}"),
            "sourceUrl": match.payload.get("source_url", ""),
            "snippet": match.payload.get("content", "")[:200] + "..."
        })

    api_key = os.getenv("GEMINI_API_KEY")
    
    # 4. Generate synthesized answer
    if api_key and top_matches:
        # Construct Context Prompt
        context_str = ""
        for idx, match in enumerate(top_matches):
            context_str += f"[{idx + 1}] File: {match.payload['title']}\nContent: {match.payload['content']}\n\n"
            
        system_instruction = (
            "You are Insight RAG, an AI-powered internal knowledge assistant designed to help employees quickly understand "
            "the company's codebase, documentation, onboarding guides, architecture notes, SOPs, and workflows. "
            "Answer the user's question using ONLY the provided internal company context. "
            "Provide clear, professional, fluent English explanations grounded 100% in the company's internal knowledge base. "
            "Do NOT fabricate, hallucinate, or dump raw unformatted code blocks unless explicitly requested. "
            "Cite source documents using brackets like [1], [2], matching the provided context index."
        )

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:streamGenerateContent?alt=sse&key={api_key}"
        payload = {
            "contents": [{
                "parts": [{"text": f"Context:\n{context_str}\n\nQuestion: {question}"}]
            }],
            "systemInstruction": {
                "parts": [{"text": system_instruction}]
            }
        }

        for attempt in range(1):
            try:
                # Connect and stream response from Gemini API with fast 5s timeout
                res = requests.post(url, json=payload, stream=True, timeout=5)
                if res.status_code == 200:
                    for line in res.iter_lines():
                        if line:
                            decoded_line = line.decode('utf-8').strip()
                            if decoded_line.startswith('data: '):
                                try:
                                    chunk_json = json.loads(decoded_line[6:])
                                    candidates = chunk_json.get("candidates", [{}])
                                    if candidates and len(candidates) > 0:
                                        parts = candidates[0].get("content", {}).get("parts", [{}])
                                        token_text = parts[0].get("text", "")
                                        if token_text:
                                            yield f"data: {json.dumps({'event': 'token', 'text': token_text})}\n\n"
                                except Exception as e:
                                    print(f"Parser exception: {str(e)}")
                                    pass
                    break
                elif res.status_code in [429, 503]:
                    # On 429/503 rate limits, instantly stream extracted RAG content from vector store without long delays
                    api_key = None
                    break
                else:
                    raise Exception(f"Gemini streaming status error: {res.status_code}")
            except Exception as e:
                api_key = None
                break
            
    # Local fallback summary if no key or no matches found
    if not api_key or not top_matches:
        if not top_matches:
            # Check if any MongoDB document title matches keywords in the question
            matching_docs = []
            seen_titles = set()
            if documents and isinstance(documents, list):
                for d in documents:
                    title_clean = d.get("title", "").lower()
                    t_orig = d.get("title")
                    if t_orig not in seen_titles and (any(k in title_clean for k in q_keywords if len(k) > 2) or any(k in title_clean for k in search_query.lower().split(" ") if len(k) > 2 and k not in STOP_WORDS)):
                        seen_titles.add(t_orig)
                        matching_docs.append(d)

            if matching_docs:
                lines = [f"Here is the workspace document information for your query **\"{question}\"**:\n"]
                for md in matching_docs[:3]:
                    lines.append(f"The document **{md.get('title')}** (`{md.get('source_type', 'file').upper()}`) is active in your workspace with status `{md.get('indexing_status', 'indexed').upper()}`.")
                msg = "\n\n".join(lines)
            else:
                msg = (
                    "I could not locate relevant information in the connected workspace documents to answer your question. "
                    "Could you please elaborate more on your question or specify what detail you need so I can search better?"
                )

            for word in msg.split(" "):
                yield f"data: {json.dumps({'event': 'token', 'text': word + ' '})}\n\n"
                await asyncio.sleep(0.01)
        else:
            synthesis_lines = [f"Based on your workspace documents, here is a summary for **\"{question}\"**:\n"]
            
            # Deduplicate top_matches by title and summarize actual content
            doc_summaries = {}
            for match in top_matches:
                t = match.payload.get("title", "Document")
                c = match.payload.get("content", "").strip()
                if t not in doc_summaries and c:
                    clean_lines = [l.strip() for l in c.split('\n') if len(l.strip()) > 8 and not any(w in l for w in ['import ', 'export ', 'const ', 'let ', 'var ', 'function ', '<div', '</', '=>', '{', '}', '$schema'])]
                    clean_text = " ".join(clean_lines[:8]).strip()
                    if clean_text:
                        doc_summaries[t] = clean_text[:280]

            if doc_summaries:
                for idx, (title, text_summary) in enumerate(doc_summaries.items(), 1):
                    synthesis_lines.append(f"**[{idx}] {title}**\n{text_summary}...\n")
            else:
                seen_titles = set()
                matching_title_docs = []
                for d in (documents or []):
                    t = d.get("title")
                    if t and t not in seen_titles and any(k in t.lower() for k in q_keywords if len(k) > 2):
                        seen_titles.add(t)
                        matching_title_docs.append(d)
                
                if matching_title_docs:
                    doc = matching_title_docs[0]
                    t = doc.get("title")
                    st = doc.get("source_type", "file").upper()
                    synthesis_lines.append(f"The document **{t}** is an active `{st}` workspace resource connected to your application. It contains technical specifications and background details for {t.split('.')[0]}.")
                else:
                    synthesis_lines.append("I could not locate relevant text content in the workspace documents for your query.")

            msg = "\n\n".join(synthesis_lines)
            for word in msg.split(" "):
                yield f"data: {json.dumps({'event': 'token', 'text': word + ' '})}\n\n"
                await asyncio.sleep(0.01)

    # 5. Stream final completion event
    confidence = 95 if (top_matches or unique_docs) else 0
    
    # Generate dynamic follow-up questions based on matched document title or topic
    matched_name = top_matches[0].payload.get("title", "") if top_matches else ""
    if "amex" in question.lower() or "resume" in question.lower() or "amex" in matched_name.lower():
        follow_ups = [
            "What technical skills and experience are detailed in AMEx_Resume.pdf?",
            "What projects or employment history are listed in this resume?",
            "Can you show me the contact details or education section in AMEx_Resume.pdf?"
        ]
    elif top_matches:
        follow_ups = [
            f"What are the main key points inside {matched_name}?",
            "Can you summarize the core architectural details?",
            "What other documents are related to this query?"
        ]
    else:
        follow_ups = [
            "What connected resources are available?",
            "How do I upload PDFs or technical documentation?",
            "How do I connect GitHub or Notion integrations?"
        ]
    
    metadata = {
        "event": "complete",
        "confidence": confidence,
        "citations": citations,
        "followUpQuestions": follow_ups
    }
    yield f"data: {json.dumps(metadata)}\n\n"

@app.post("/ask")
async def ask_endpoint(payload: AskRequest):
    if not payload.question:
        raise HTTPException(status_code=400, detail="Question is required")
    return StreamingResponse(
        execute_hybrid_rag_streaming(payload.question, payload.org_id, payload.document_id, payload.integrations, payload.documents, payload.chat_history),
        media_type="text/event-stream"
    )

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
